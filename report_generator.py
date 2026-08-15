"""
report_generator.py
Đọc file Excel doanh thu thô (định dạng cột giống Bách Hóa Xanh xuất ra)
và tạo báo cáo gọn: Doanh thu / Lượt bill / Giá trị bill theo Offline - Online - Tổng.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

FONT_NAME = "Arial"

# Tên cột trong file thô (phải khớp với file BHX xuất ra)
COL = {
    "ngay": "Ngày",
    "ma_st": "Mã siêu thị",
    "ten_st": "Tên siêu thị",
    "dt_offline": "Doanh thu offline",
    "dt_online": "Doanh thu Online",
    "bill_offline": "Tổng số bill",
    "bill_online": "Tổng số bill online",
    "tinh": "Tỉnh/TP",
}


def _find_col_index(header_row, name):
    for i, cell in enumerate(header_row, start=1):
        if str(cell.value).strip() == name:
            return i
    return None


def read_raw_data(input_path):
    """Đọc dòng dữ liệu đầu tiên (dòng 2) của sheet đầu tiên trong file thô."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]

    header_row = list(ws[1])
    col_idx = {key: _find_col_index(header_row, label) for key, label in COL.items()}

    missing = [k for k, v in col_idx.items() if v is None]
    if missing:
        raise ValueError(
            f"Không tìm thấy các cột: {[COL[m] for m in missing]}. "
            "Kiểm tra lại file có đúng định dạng doanh thu theo siêu thị không."
        )

    row = 2
    data = {}
    for key, idx in col_idx.items():
        data[key] = ws.cell(row=row, column=idx).value

    return data


def build_report(data, output_path):
    """Tạo file Excel báo cáo gọn (Doanh thu / Lượt bill / Giá trị bill)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo doanh thu"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    title_font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    sub_font = Font(name=FONT_NAME, size=11, color="595959")
    label_font = Font(name=FONT_NAME, size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    ws['A1'] = "BÁO CÁO DOANH THU NGÀY"
    ws['A1'].font = title_font

    ws.merge_cells('A2:D2')
    ws['A2'] = f"{data['ten_st']}"
    ws['A2'].font = sub_font

    ngay = data['ngay']
    if isinstance(ngay, datetime):
        ngay_str = ngay.strftime("%d/%m/%Y")
    else:
        ngay_str = str(ngay)
    ws.merge_cells('A3:D3')
    ws['A3'] = ngay_str
    ws['A3'].font = sub_font

    r = 5
    headers2 = ["Kênh", "Doanh thu", "Lượt bill", "Giá trị bill"]
    for i, h in enumerate(headers2):
        cell = ws.cell(row=r, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    r += 1

    dt_offline = float(data['dt_offline'] or 0)
    dt_online = float(data['dt_online'] or 0)
    bill_offline = float(data['bill_offline'] or 0)
    bill_online = float(data['bill_online'] or 0)

    rows = [
        ("Offline", dt_offline, bill_offline),
        ("Online", dt_online, bill_online),
    ]

    channel_start = r
    for name, dt, bill in rows:
        gia_tri = dt / bill if bill else 0
        ws.cell(row=r, column=1, value=name).font = label_font
        ws.cell(row=r, column=2, value=dt).font = label_font
        ws.cell(row=r, column=2).number_format = '#,##0" đ"'
        ws.cell(row=r, column=3, value=bill).font = label_font
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=4, value=gia_tri).font = label_font
        ws.cell(row=r, column=4).number_format = '#,##0" đ"'
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = border
        r += 1
    channel_end = r - 1

    total_dt = dt_offline + dt_online
    total_bill = bill_offline + bill_online
    total_gia_tri = total_dt / total_bill if total_bill else 0

    ws.cell(row=r, column=1, value="Tổng").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=2, value=total_dt).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=2).number_format = '#,##0" đ"'
    ws.cell(row=r, column=3, value=total_bill).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=3).number_format = '#,##0'
    ws.cell(row=r, column=4, value=total_gia_tri).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=4).number_format = '#,##0" đ"'
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    wb.save(output_path)

    return {
        "ten_st": data['ten_st'],
        "ngay_str": ngay_str,
        "dt_offline": dt_offline,
        "dt_online": dt_online,
        "total_dt": total_dt,
        "bill_offline": bill_offline,
        "bill_online": bill_online,
        "total_bill": total_bill,
        "gia_tri_offline": dt_offline / bill_offline if bill_offline else 0,
        "gia_tri_online": dt_online / bill_online if bill_online else 0,
        "total_gia_tri": total_gia_tri,
    }


def build_text_summary(summary):
    """Tạo tin nhắn tóm tắt gửi trong LINE."""
    def fmt(n):
        return f"{n:,.0f}".replace(",", ".")

    lines = [
        f"📊 BÁO CÁO DOANH THU - {summary['ngay_str']}",
        f"{summary['ten_st']}",
        "",
        "▪ Offline",
        f"  Doanh thu: {fmt(summary['dt_offline'])} đ",
        f"  Lượt bill: {fmt(summary['bill_offline'])}",
        f"  Giá trị bill: {fmt(summary['gia_tri_offline'])} đ",
        "",
        "▪ Online",
        f"  Doanh thu: {fmt(summary['dt_online'])} đ",
        f"  Lượt bill: {fmt(summary['bill_online'])}",
        f"  Giá trị bill: {fmt(summary['gia_tri_online'])} đ",
        "",
        "▪ TỔNG",
        f"  Doanh thu: {fmt(summary['total_dt'])} đ",
        f"  Lượt bill: {fmt(summary['total_bill'])}",
        f"  Giá trị bill: {fmt(summary['total_gia_tri'])} đ",
    ]
    return "\n".join(lines)
