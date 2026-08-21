"""
excel_reader.py - Đọc 3 loại file Excel mà bot nhận:

1) File "doanh thu theo siêu thị" (nhiều dòng, mỗi dòng 1 siêu thị / 1 ngày)
   -> read_all_rows()  (giữ nguyên logic cũ, không đổi)

2) File "doanh thu chi tiết" theo ngành hàng / sản phẩm (nhiều dòng, mỗi dòng
   1 sản phẩm bán ra trong ngày, cùng 1 siêu thị)
   -> read_category_rows()  (phục vụ báo cáo "MỤC TIÊU KHUYẾN MÃI":
      Nấm / Bánh trung thu / Trà C2)

3) File "BC tồn theo model" (tồn kho từng sản phẩm tại siêu thị)
   -> read_stock_rows()  (mới thêm, phục vụ tính "% bán trên tồn" cho
      Bánh trung thu và Trà C2 trong báo cáo MỤC TIÊU KHUYẾN MÃI)

detect_file_type() dùng để tự động phân biệt 3 loại file khi người dùng gửi
file vào bot, không cần người dùng khai báo loại file.
"""
import openpyxl
from datetime import datetime

# ---------------------------------------------------------------------------
# LOẠI 1: FILE DOANH THU THEO SIÊU THỊ (giữ nguyên như code cũ)
# ---------------------------------------------------------------------------
COL = {
    "ngay": "Ngày",
    "ma_st": "Mã siêu thị",
    "ten_st": "Tên siêu thị",
    # Dùng số liệu CHƯA VAT theo lựa chọn của anh
    "dt_offline": "Doanh thu offline",
    "dt_online": "Doanh thu Online",
    "bill_offline": "Tổng số bill",
    "bill_online": "Tổng số bill online",
    "tinh": "Tỉnh/TP",
}

# Cột đặc trưng của file "chi tiết ngành hàng" (per-product)
CATEGORY_COL = {
    "ngay": "Ngày xuất",
    "ten_st": "Tên siêu thị",
    "ma_sp": "Mã sản phẩm",
    "ten_sp": "Tên sản phẩm",
    "nganh_hang": "Ngành hàng",
    "nhom_hang": "Nhóm hàng",
    "ten_hang": "Tên Hãng",
    "don_vi": "Đơn vị",
    "sl_thuc_xuat": "SL thực xuất",
    "thanh_tien": "Thành tiền phải thu khách hàng (chưa VAT)",
    "hinh_thuc_xuat": "Tên hình thức xuất bán",
}

# Cột đặc trưng của file "BC tồn theo model" (tồn kho)
STOCK_COL = {
    "model": "Model",
    "ma_sp": "Mã sản phẩm",
    "ten_st": "Tên siêu thị",
    "ton_kho": "Tồn kho siêu thị",
    "don_vi": "Đơn vị",
}


def _header_row(ws):
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def detect_file_type(input_path):
    """Trả về 'revenue', 'category', 'stock', hoặc None nếu không nhận diện được."""
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    headers = set(_header_row(ws))
    wb.close()

    if {"Doanh thu offline", "Doanh thu Online", "Mã siêu thị"}.issubset(headers):
        return "revenue"
    if {"Tên sản phẩm", "Ngành hàng", "Nhóm hàng"}.issubset(headers):
        return "category"
    if {"Tồn kho siêu thị", "Mã sản phẩm", "Model"}.issubset(headers):
        return "stock"
    return None


def _find_col_index(header_row, name):
    for i, cell in enumerate(header_row, start=1):
        if str(cell.value).strip() == name:
            return i
    return None


def read_all_rows(input_path):
    """Đọc toàn bộ dòng dữ liệu trong sheet đầu tiên (file doanh thu theo siêu thị).
    Trả về (ngay_str, list_of_dict) — mỗi dict là 1 siêu thị."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    header_row = list(ws[1])
    col_idx = {key: _find_col_index(header_row, label) for key, label in COL.items()}
    missing = [k for k, v in col_idx.items() if v is None]
    if missing:
        raise ValueError(
            f"Không tìm thấy các cột: {[COL[m] for m in missing]}. "
            "Kiểm tra lại file có đúng định dạng doanh thu theo siêu thị không."
        )
    rows = []
    for r in range(2, ws.max_row + 1):
        ma_st = ws.cell(row=r, column=col_idx["ma_st"]).value
        if ma_st is None:
            continue
        ngay = ws.cell(row=r, column=col_idx["ngay"]).value
        if isinstance(ngay, datetime):
            ngay_str = ngay.strftime("%Y-%m-%d")
        else:
            ngay_str = str(ngay)[:10]
        rows.append({
            "ngay": ngay_str,
            "ma_st": str(ma_st),
            "ten_st": ws.cell(row=r, column=col_idx["ten_st"]).value,
            "tinh": ws.cell(row=r, column=col_idx["tinh"]).value,
            "dt_offline": float(ws.cell(row=r, column=col_idx["dt_offline"]).value or 0),
            "dt_online": float(ws.cell(row=r, column=col_idx["dt_online"]).value or 0),
            "bill_offline": float(ws.cell(row=r, column=col_idx["bill_offline"]).value or 0),
            "bill_online": float(ws.cell(row=r, column=col_idx["bill_online"]).value or 0),
        })
    if not rows:
        raise ValueError("Không tìm thấy dòng dữ liệu nào trong file.")
    return rows[0]["ngay"], rows


# ---------------------------------------------------------------------------
# LOẠI 2: FILE CHI TIẾT NGÀNH HÀNG (Nấm / Bánh trung thu / Trà C2)
# ---------------------------------------------------------------------------

# Các sản phẩm C2 cần theo dõi, quy đổi ra "chai".
# key = tên hiển thị, value = từ khoá để nhận diện trong "Tên sản phẩm"
C2_TARGETS = [
    ("Chanh tuyết bạc hà", "TUYẾT"),
    ("Trà xanh hương chanh 360ml", "HƯƠNG CHANH"),
    ("Trà hồng vải", "HỒNG VẢI"),
    ("Trà vải", "VẢI"),
    ("Trà đen dâu anh đào", "DÂU ANH ĐÀO"),
    ("Sâm Cúc", "SÂM"),
    ("Trà đen tắc", "TẮC"),
]

UNIT_TO_CHAI = {
    "THÙNG": 24,
    "LỐC": 6,
    "CHAI": 1,
}


def _unit_multiplier(don_vi):
    if not don_vi:
        return 1
    key = str(don_vi).strip().upper()
    return UNIT_TO_CHAI.get(key, 1)


# 5 nhóm ngành hàng gộp theo yêu cầu, hiển thị trong bảng "DOANH THU THEO NGÀNH HÀNG"
# của thẻ báo cáo doanh thu. key = tên hiển thị, value = danh sách tên "Ngành hàng"
# gốc trong file cần cộng gộp lại.
NGANH_HANG_GROUPS = [
    ("Bia Nước", ["Bia Các Loại", "Thức uống giải khát các loại"]),
    ("Sữa - Đông Mát", ["Sữa - Thức uống bổ dưỡng các loại", "Thực phẩm đông lạnh - Hàng mát các loại"]),
    ("Thịt Gia Cầm", ["Thịt gia cầm gia súc các loại"]),
    ("Trái Cây - Rau Củ", ["Trái Cây Các Loại", "Rau Củ Các Loại"]),
    ("Thủy Hải Sản", ["Thủy Hải Sản Các Loại"]),
]


def read_category_rows(input_path, filter_date=None):
    """Đọc file chi tiết ngành hàng, trả về dict tổng hợp:
    {
        "ngay": "2026-08-15",
        "ten_st": "BHX_STR_CLD - Thửa 1289 An Nghiệp",
        "nam": {"doanh_thu": 798095},
        "banh_trung_thu": {
            "items": [{"ten": "...", "sl": 1, "thanh_tien": 78704}, ...],
            "tong_sl": 5,
            "tong_tien": 307407,
        },
        "c2": {
            "items": [{"ten": "...", "chai": 494, "thanh_tien": 2872727}, ...],
            "tong_chai": 694,
            "tong_tien": 4054764,
        },
    }

    filter_date: nếu file trải nhiều ngày (vd dùng chung cho báo cáo thưởng),
    truyền vào "YYYY-MM-DD" để CHỈ lấy đúng ngày đó (dùng cho MTKM — luôn là
    dữ liệu 1 ngày). Để None nếu file vốn đã chỉ có 1 ngày.
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    header_row = list(ws[1])
    col_idx = {key: _find_col_index(header_row, label) for key, label in CATEGORY_COL.items()}
    missing = [k for k, v in col_idx.items() if v is None]
    if missing:
        raise ValueError(
            f"Không tìm thấy các cột: {[CATEGORY_COL[m] for m in missing]}. "
            "Kiểm tra lại file có đúng định dạng chi tiết ngành hàng không."
        )

    def cell(r, key):
        return ws.cell(row=r, column=col_idx[key]).value

    ngay_str = None
    ten_st = None

    nam_total = 0.0

    # gộp theo tên sản phẩm để không lặp dòng trùng
    btt_map = {}   # ten_sp -> {sl, thanh_tien, ma_sp_set}
    c2_map = {label: {"chai": 0.0, "thanh_tien": 0.0, "sl_raw": 0.0, "ma_sp_set": set()} for label, _ in C2_TARGETS}
    nganh_hang_map = {}  # ten_nganh_hang -> thanh_tien (tổng doanh thu toàn bộ ngành hàng)

    for r in range(2, ws.max_row + 1):
        ten_sp = cell(r, "ten_sp")
        if ten_sp is None:
            continue
        ten_sp_str = str(ten_sp).strip()
        ten_sp_upper = ten_sp_str.upper()

        ngay_val = cell(r, "ngay")
        row_ngay_str = None
        if isinstance(ngay_val, datetime):
            row_ngay_str = ngay_val.strftime("%Y-%m-%d")
        elif ngay_val:
            row_ngay_str = str(ngay_val)[:10]

        if filter_date is not None:
            if row_ngay_str != filter_date:
                continue
            ngay_str = filter_date
        elif ngay_str is None and row_ngay_str:
            ngay_str = row_ngay_str

        if ten_st is None:
            v = cell(r, "ten_st")
            if v:
                ten_st = str(v).strip()

        sl_thuc_xuat = float(cell(r, "sl_thuc_xuat") or 0)
        thanh_tien = float(cell(r, "thanh_tien") or 0)
        nhom_hang = str(cell(r, "nhom_hang") or "").strip()
        ten_hang = str(cell(r, "ten_hang") or "").strip().upper()
        don_vi = cell(r, "don_vi")
        hinh_thuc = str(cell(r, "hinh_thuc_xuat") or "").strip().upper()
        ma_sp_val = cell(r, "ma_sp")
        ma_sp = str(ma_sp_val).strip() if ma_sp_val is not None else ""

        # ---- NẤM ----
        if nhom_hang == "Nấm Các Loại":
            nam_total += thanh_tien

        # ---- TỔNG DOANH THU THEO NGÀNH HÀNG (tất cả ngành hàng) ----
        nganh_hang = str(cell(r, "nganh_hang") or "").strip()
        if nganh_hang:
            nganh_hang_map[nganh_hang] = nganh_hang_map.get(nganh_hang, 0.0) + thanh_tien

        # ---- BÁNH TRUNG THU ----
        if "TRUNG THU" in ten_sp_upper:
            # bỏ qua hàng tặng / khuyến mãi, chỉ tính hàng bán thực tế
            if "TẶNG" not in hinh_thuc and sl_thuc_xuat > 0:
                entry = btt_map.setdefault(ten_sp_str, {"sl": 0.0, "thanh_tien": 0.0, "ma_sp_set": set()})
                entry["sl"] += sl_thuc_xuat
                entry["thanh_tien"] += thanh_tien
                if ma_sp:
                    entry["ma_sp_set"].add(ma_sp)

        # ---- TRÀ C2 ----
        if ten_hang == "C2" or "C2" in ten_sp_upper:
            if sl_thuc_xuat > 0:
                for label, keyword in C2_TARGETS:
                    if keyword in ten_sp_upper:
                        mult = _unit_multiplier(don_vi)
                        c2_map[label]["chai"] += sl_thuc_xuat * mult
                        c2_map[label]["thanh_tien"] += thanh_tien
                        c2_map[label]["sl_raw"] += sl_thuc_xuat
                        if ma_sp:
                            c2_map[label]["ma_sp_set"].add(ma_sp)
                        break

    if ngay_str is None:
        raise ValueError("Không tìm thấy dữ liệu ngày trong file.")

    btt_items = [
        {"ten": ten, "sl": v["sl"], "thanh_tien": v["thanh_tien"], "ma_sp": sorted(v["ma_sp_set"])}
        for ten, v in sorted(btt_map.items())
        if v["sl"] > 0
    ]
    c2_items = [
        {"ten": label, "chai": v["chai"], "thanh_tien": v["thanh_tien"],
         "sl_raw": v["sl_raw"], "ma_sp": sorted(v["ma_sp_set"])}
        for label, v in c2_map.items()
        if v["chai"] > 0
    ]

    nganh_hang_items = [
        {"ten": ten, "thanh_tien": tt}
        for ten, tt in sorted(nganh_hang_map.items(), key=lambda x: -x[1])
    ]

    return {
        "ngay": ngay_str,
        "ten_st": ten_st or "—",
        "nam": {"doanh_thu": nam_total},
        "banh_trung_thu": {
            "items": btt_items,
            "tong_sl": sum(i["sl"] for i in btt_items),
            "tong_tien": sum(i["thanh_tien"] for i in btt_items),
        },
        "c2": {
            "items": c2_items,
            "tong_chai": sum(i["chai"] for i in c2_items),
            "tong_tien": sum(i["thanh_tien"] for i in c2_items),
        },
        "nganh_hang": {
            "items": nganh_hang_items,
            "tong_tien": sum(i["thanh_tien"] for i in nganh_hang_items),
        },
    }


# ---------------------------------------------------------------------------
# LOẠI 3: FILE TỒN KHO (BC tồn theo model)
# ---------------------------------------------------------------------------

def count_distinct_dates(input_path):
    """Đếm nhanh số ngày khác nhau trong cột 'Ngày xuất' của file category,
    dùng để phân biệt file 1 ngày (MTKM) và file nhiều ngày (báo cáo thưởng)."""
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    idx = header.index("Ngày xuất")
    dates = set()
    for row in rows_iter:
        v = row[idx]
        if v:
            d = v.date() if isinstance(v, datetime) else str(v)[:10]
            dates.add(d)
    wb.close()
    return len(dates)


def read_stock_rows(input_path):
    """Đọc file "BC tồn theo model", trả về dict:
    {
        "ten_st": "BHX_STR_CLD - Thửa 1289 An Nghiệp",
        "ton_kho_map": {"8888077102092": 12.0, ...},   # mã sản phẩm (đã strip) -> tồn kho
        "rows": [{"model": "...", "ton_kho": 12.0}, ...],  # dùng để so khớp theo từ khóa (vd Trà C2)
    }
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    header_row = list(ws[1])
    col_idx = {key: _find_col_index(header_row, label) for key, label in STOCK_COL.items()}
    missing = [k for k, v in col_idx.items() if v is None]
    if missing:
        raise ValueError(
            f"Không tìm thấy các cột: {[STOCK_COL[m] for m in missing]}. "
            "Kiểm tra lại file có đúng định dạng tồn kho (BC tồn theo model) không."
        )

    def cell(r, key):
        return ws.cell(row=r, column=col_idx[key]).value

    ten_st = None
    ton_kho_map = {}
    rows = []

    for r in range(2, ws.max_row + 1):
        ma_sp_val = cell(r, "ma_sp")
        if ma_sp_val is None:
            continue
        ma_sp = str(ma_sp_val).strip()
        if not ma_sp:
            continue

        if ten_st is None:
            v = cell(r, "ten_st")
            if v:
                ten_st = str(v).strip()

        ton_kho = float(cell(r, "ton_kho") or 0)
        ton_kho_map[ma_sp] = ton_kho_map.get(ma_sp, 0.0) + ton_kho

        model_val = cell(r, "model")
        model_str = str(model_val).strip() if model_val is not None else ""
        rows.append({"model": model_str, "ton_kho": ton_kho})

    if not ton_kho_map:
        raise ValueError("Không tìm thấy dòng dữ liệu tồn kho nào trong file.")

    return {"ten_st": ten_st or "—", "ton_kho_map": ton_kho_map, "rows": rows}


def attach_stock_percentage(category_payload, stock_data):
    """Ghép % bán/nhập vào từng sản phẩm Bánh trung thu / Trà C2 của báo cáo
    ngành hàng. Công thức: % = SL bán (đã quy đổi cùng đơn vị) / Tồn kho hiện tại.

    - Bánh trung thu: so khớp theo mã sản phẩm (đơn vị "Cái" đồng nhất 2 file).
    - Trà C2: so khớp theo TỪ KHÓA tên sản phẩm (giống cách nhận diện C2_TARGETS),
      vì mã sản phẩm giữa 2 file không đồng nhất; số bán dùng "chai" đã quy đổi
      để cùng đơn vị với tồn kho (tồn kho file luôn tính theo Chai).

    Không sửa category_payload gốc — trả về bản sao đã ghép thêm "pct_ban_nhap".
    """
    import copy
    payload = copy.deepcopy(category_payload)

    ton_kho_map = stock_data.get("ton_kho_map", {})
    stock_rows = stock_data.get("rows", [])

    def _pct_btt(ma_sp_list, sl_ban):
        ton_kho_tong = sum(ton_kho_map.get(ma, 0.0) for ma in ma_sp_list)
        if ton_kho_tong <= 0:
            return None
        return sl_ban / ton_kho_tong * 100

    def _ton_kho_c2_theo_nhan():
        """Tính tồn kho cho từng nhãn C2, dùng logic khớp từ khóa ĐẦU TIÊN
        (giống hệt cách phân loại lúc đọc file doanh thu chi tiết) để một sản
        phẩm không bị tính trùng vào 2 nhãn cùng lúc (vd "chanh tuyết bạc hà"
        chứa cả từ khóa "HƯƠNG CHANH" lẫn "TUYẾT")."""
        result = {label: 0.0 for label, _ in C2_TARGETS}
        for row in stock_rows:
            m = row["model"].upper()
            if "C2" not in m:
                continue
            for label, keyword in C2_TARGETS:
                if keyword in m:
                    result[label] += row["ton_kho"]
                    break
        return result

    ton_kho_c2 = _ton_kho_c2_theo_nhan()

    for item in payload["banh_trung_thu"]["items"]:
        item["pct_ban_nhap"] = _pct_btt(item.get("ma_sp", []), item["sl"])

    for item in payload["c2"]["items"]:
        ton_kho_tong = ton_kho_c2.get(item["ten"], 0.0)
        item["pct_ban_nhap"] = (item["chai"] / ton_kho_tong * 100) if ton_kho_tong > 0 else None

    return payload


# ---------------------------------------------------------------------------
# BÁO CÁO THƯỞNG (lệnh "TD" / "THƯỞNG") — FRESH (thịt heo/gà nhập khẩu)
# + FMCG (Sữa-Kem-Đông-Mát, Bia-Nước, Trà C2)
#
# Dùng chung file "doanh thu chi tiết" như MTKM, nhưng có thể trải NHIỀU NGÀY
# trong 1 file (ví dụ từ đầu tháng đến hiện tại). Base của từng chương trình
# lấy theo mặc định đã tính sẵn từ dữ liệu thực tế các tháng trước — không
# cần người dùng tự nhập.
# ---------------------------------------------------------------------------

import calendar
from datetime import date as _date

# ----- Base mặc định (đã tính sẵn từ dữ liệu tháng 5, 6, 7/2026 thực tế) -----
THUONG_BASE = {
    "fresh_kg": 930.6,                 # TB tháng 5-6/2026 (thịt heo/gà nhập khẩu)
    "skdm_revenue": 426_257_994.0,     # Doanh thu Sữa-Kem-Đông-Mát tháng 7/2026
    "bianuoc_revenue": 175_070_233.0,  # Doanh thu Bia-Nước tháng 6/2026
}

SKDM_NGANH_HANG = {
    "Sữa - Thức uống bổ dưỡng các loại",
    "Thực phẩm đông lạnh - Hàng mát các loại",
    "Sản Phẩm Từ Sữa - Bảo Quản Mát",
    "Kem các loại",
}
BIANUOC_NGANH_HANG = {"Bia Các Loại", "Thức uống giải khát các loại"}

# (tên hiển thị, các từ khóa BẮT BUỘC cùng xuất hiện trong tên sản phẩm, đơn giá thưởng/chai)
THUONG_C2_TARGETS = [
    ("Freeze Chanh Tuyết", ["FREEZE", "TUYẾT"], 500),
    ("Freeze Dâu Anh Đào", ["FREEZE", "DÂU ANH ĐÀO"], 500),
    ("Tắc", ["TẮC"], 500),
    ("Freeze Vải", ["FREEZE", "VẢI"], 500),
    ("Sâm Cúc", ["SÂM"], 500),
]
THUONG_OLONG_KEYWORDS = ["OOLONG", "CHAI 1L"]
THUONG_OLONG_RATE = 1000


def _muc_thuong_fresh(pct):
    if pct >= 12:
        return 1000, "Mức cao nhất (≥12%)"
    if pct >= 8:
        return 800, "Mức 2 (≥8%)"
    if pct >= 5:
        return 300, "Mức 1 (≥5%)"
    return 0, "Chưa đạt mốc thưởng (<5%)"


def _muc_thuong_skdm(pct):
    if pct >= 10:
        return 0.005, "Mức 2 (≥10%)"
    if pct >= 5:
        return 0.0025, "Mức 1 (≥5%)"
    return 0.0, "Chưa đạt mốc thưởng (<5%)"


def _muc_thuong_bianuoc(pct):
    if pct >= 25:
        return 750_000, "Mức 3 (≥25%, cao nhất)"
    if pct >= 15:
        return 400_000, "Mức 2 (≥15%)"
    if pct >= 5:
        return 250_000, "Mức 1 (≥5%)"
    return 0, "Chưa đạt mốc thưởng (<5%)"


def read_thuong_period_rows(input_path):
    """Đọc file "doanh thu chi tiết" (có thể trải nhiều ngày, vd 01/08 -> hiện tại).
    Trả về dict payload đầy đủ số liệu FRESH + FMCG, đã áp base mặc định và
    quy đổi theo nhịp độ (số ngày đã qua -> ước cả tháng)."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    header_row = list(ws[1])

    col_map = {
        "ngay": "Ngày xuất", "ten_st": "Tên siêu thị", "ten_sp": "Tên sản phẩm",
        "nganh_hang": "Ngành hàng", "nhom_hang": "Nhóm hàng", "ten_hang": "Tên Hãng",
        "don_vi": "Đơn vị", "sl_thuc_xuat": "SL thực xuất",
        "thanh_tien": "Thành tiền phải thu khách hàng (chưa VAT)",
        "nh_phan_tich": "Ngành hàng - Phân tích",
    }
    col_idx = {k: _find_col_index(header_row, v) for k, v in col_map.items()}
    missing = [k for k, v in col_idx.items() if v is None]
    if missing:
        raise ValueError(f"Không tìm thấy các cột: {[col_map[m] for m in missing]}.")

    def cell(r, key):
        return ws.cell(row=r, column=col_idx[key]).value

    ten_st = None
    ngay_min = None
    ngay_max = None

    fresh_kg = 0.0
    fresh_tien = 0.0
    skdm_tien = 0.0
    bianuoc_tien = 0.0
    c2_chai = {label: 0.0 for label, _, _ in THUONG_C2_TARGETS}
    olong_chai = 0.0

    for r in range(2, ws.max_row + 1):
        ten_sp = cell(r, "ten_sp")
        if ten_sp is None:
            continue
        ten_sp_upper = str(ten_sp).strip().upper()

        ngay_val = cell(r, "ngay")
        if isinstance(ngay_val, datetime):
            d = ngay_val.date()
        elif ngay_val:
            try:
                d = datetime.strptime(str(ngay_val)[:10], "%Y-%m-%d").date()
            except Exception:
                d = None
        else:
            d = None
        if d:
            ngay_min = d if ngay_min is None or d < ngay_min else ngay_min
            ngay_max = d if ngay_max is None or d > ngay_max else ngay_max

        if ten_st is None:
            v = cell(r, "ten_st")
            if v:
                ten_st = str(v).strip()

        thanh_tien = float(cell(r, "thanh_tien") or 0)
        sl = float(cell(r, "sl_thuc_xuat") or 0)
        nhom_hang = str(cell(r, "nhom_hang") or "").strip()
        nganh_hang = str(cell(r, "nganh_hang") or "").strip()
        nh_pt = str(cell(r, "nh_phan_tich") or "").strip()
        don_vi = str(cell(r, "don_vi") or "").strip().upper()
        ten_hang = str(cell(r, "ten_hang") or "").strip().upper()

        # ---- FRESH: thịt heo/gà nhập khẩu ----
        if nhom_hang in ("Thịt Heo Các Loại", "Thịt Gia Cầm Các Loại") and nh_pt == "Thịt Nhập Khẩu":
            fresh_tien += thanh_tien
            if don_vi == "KG":
                fresh_kg += sl

        # ---- SKDM ----
        if nganh_hang in SKDM_NGANH_HANG:
            skdm_tien += thanh_tien

        # ---- BIA-NUOC ----
        if nganh_hang in BIANUOC_NGANH_HANG:
            bianuoc_tien += thanh_tien

        # ---- TRÀ C2 (thưởng theo sản phẩm) ----
        if (ten_hang == "C2" or "C2" in ten_sp_upper) and sl > 0:
            mult = _unit_multiplier(don_vi)
            chai = sl * mult
            for label, keywords, _rate in THUONG_C2_TARGETS:
                if all(kw in ten_sp_upper for kw in keywords):
                    c2_chai[label] += chai
                    break

        # ---- OOLONG 1L (không thuộc nhãn C2, kiểm tra riêng) ----
        if sl > 0 and all(kw in ten_sp_upper for kw in THUONG_OLONG_KEYWORDS):
            mult = _unit_multiplier(don_vi)
            olong_chai += sl * mult

    if ngay_min is None:
        raise ValueError("Không tìm thấy dữ liệu ngày trong file.")

    so_ngay_da_qua = (ngay_max - ngay_min).days + 1
    so_ngay_ca_thang = calendar.monthrange(ngay_max.year, ngay_max.month)[1]
    he_so_quy_doi = so_ngay_ca_thang / so_ngay_da_qua if so_ngay_da_qua else 1.0

    # ----- FRESH -----
    fresh_proj_kg = fresh_kg * he_so_quy_doi
    fresh_pct = (fresh_proj_kg / THUONG_BASE["fresh_kg"] - 1) * 100 if THUONG_BASE["fresh_kg"] else 0
    fresh_rate, fresh_muc = _muc_thuong_fresh(fresh_pct)
    fresh_thuong = fresh_proj_kg * fresh_rate

    # ----- SKDM -----
    skdm_proj = skdm_tien * he_so_quy_doi
    skdm_pct = (skdm_proj / THUONG_BASE["skdm_revenue"] - 1) * 100 if THUONG_BASE["skdm_revenue"] else 0
    skdm_rate, skdm_muc = _muc_thuong_skdm(skdm_pct)
    skdm_thuong = skdm_proj * skdm_rate

    # ----- BIA-NUOC -----
    bianuoc_proj = bianuoc_tien * he_so_quy_doi
    bianuoc_pct = (bianuoc_proj / THUONG_BASE["bianuoc_revenue"] - 1) * 100 if THUONG_BASE["bianuoc_revenue"] else 0
    bianuoc_rate, bianuoc_muc = _muc_thuong_bianuoc(bianuoc_pct)
    bianuoc_thuong = bianuoc_rate  # cố định theo mức, không nhân doanh thu

    # ----- C2 -----
    c2_items = []
    tong_c2_thuong_mtd = 0.0
    tong_c2_thuong_proj = 0.0
    for label, _keywords, rate in THUONG_C2_TARGETS:
        chai_mtd = c2_chai[label]
        chai_proj = chai_mtd * he_so_quy_doi
        thuong_mtd = chai_mtd * rate
        thuong_proj = chai_proj * rate
        tong_c2_thuong_mtd += thuong_mtd
        tong_c2_thuong_proj += thuong_proj
        c2_items.append({
            "ten": label, "chai_mtd": chai_mtd, "chai_proj": chai_proj,
            "rate": rate, "thuong_mtd": thuong_mtd, "thuong_proj": thuong_proj,
        })
    olong_proj = olong_chai * he_so_quy_doi
    olong_thuong_mtd = olong_chai * THUONG_OLONG_RATE
    olong_thuong_proj = olong_proj * THUONG_OLONG_RATE
    tong_c2_thuong_mtd += olong_thuong_mtd
    tong_c2_thuong_proj += olong_thuong_proj
    c2_items.append({
        "ten": "Olong 1L các loại", "chai_mtd": olong_chai, "chai_proj": olong_proj,
        "rate": THUONG_OLONG_RATE, "thuong_mtd": olong_thuong_mtd, "thuong_proj": olong_thuong_proj,
    })

    tong_thuong_du_kien = fresh_thuong + skdm_thuong + bianuoc_thuong + tong_c2_thuong_proj

    return {
        "ten_st": ten_st or "—",
        "ngay_bat_dau": ngay_min.strftime("%Y-%m-%d"),
        "ngay_ket_thuc": ngay_max.strftime("%Y-%m-%d"),
        "so_ngay_da_qua": so_ngay_da_qua,
        "so_ngay_ca_thang": so_ngay_ca_thang,
        "fresh": {
            "base_kg": THUONG_BASE["fresh_kg"], "thuc_te_kg": fresh_kg, "thuc_te_tien": fresh_tien,
            "du_kien_kg": fresh_proj_kg, "pct": fresh_pct, "rate": fresh_rate, "muc": fresh_muc,
            "thuong_du_kien": fresh_thuong,
        },
        "skdm": {
            "base_tien": THUONG_BASE["skdm_revenue"], "thuc_te_tien": skdm_tien,
            "du_kien_tien": skdm_proj, "pct": skdm_pct, "rate": skdm_rate, "muc": skdm_muc,
            "thuong_du_kien": skdm_thuong,
        },
        "bianuoc": {
            "base_tien": THUONG_BASE["bianuoc_revenue"], "thuc_te_tien": bianuoc_tien,
            "du_kien_tien": bianuoc_proj, "pct": bianuoc_pct, "rate": bianuoc_rate, "muc": bianuoc_muc,
            "thuong_du_kien": bianuoc_thuong,
        },
        "c2": {
            "items": c2_items,
            "tong_thuong_mtd": tong_c2_thuong_mtd,
            "tong_thuong_du_kien": tong_c2_thuong_proj,
        },
        "tong_thuong_du_kien": tong_thuong_du_kien,
    }


# ---------------------------------------------------------------------------
# LỊCH HỖ TRỢ SIÊU THỊ KHÁC (Ngày | Tên | Ca làm) - mới thêm
# ---------------------------------------------------------------------------

SCHEDULE_COL_NAMES = {"ngay": "Ngày", "ten": "Tên", "ca": "Ca làm"}


def is_schedule_file(input_path):
    """Kiểm tra nhanh xem file có phải file 'lịch hỗ trợ' (Ngày/Tên/Ca làm) không."""
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    headers = set(_header_row(ws))
    wb.close()
    return {"Ngày", "Tên", "Ca làm"}.issubset(headers)


def read_schedule_rows(input_path, default_year=None):
    """Đọc file lịch hỗ trợ (cột Ngày | Tên | Ca làm). Trả về list dict
    {"ngay": "YYYY-MM-DD", "ten": "...", "ca": "..."} — bỏ qua dòng trống."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    header_row = list(ws[1])
    col_idx = {k: _find_col_index(header_row, v) for k, v in SCHEDULE_COL_NAMES.items()}
    missing = [k for k, v in col_idx.items() if v is None]
    if missing:
        raise ValueError(f"Không tìm thấy các cột: {[SCHEDULE_COL_NAMES[m] for m in missing]}.")

    if default_year is None:
        default_year = datetime.now().year

    rows = []
    for r in range(2, ws.max_row + 1):
        ten = ws.cell(row=r, column=col_idx["ten"]).value
        ngay_val = ws.cell(row=r, column=col_idx["ngay"]).value
        if not ten or not ngay_val:
            continue
        if isinstance(ngay_val, datetime):
            ngay_str = ngay_val.strftime("%Y-%m-%d")
        else:
            s = str(ngay_val).strip()
            parts = s.replace("-", "/").split("/")
            if len(parts) == 2:
                day, month = int(parts[0]), int(parts[1])
                ngay_str = f"{default_year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 3:
                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                if year < 100:
                    year += 2000
                ngay_str = f"{year:04d}-{month:02d}-{day:02d}"
            else:
                continue
        ca_val = ws.cell(row=r, column=col_idx["ca"]).value
        rows.append({"ngay": ngay_str, "ten": str(ten).strip(), "ca": str(ca_val).strip() if ca_val else ""})

    if not rows:
        raise ValueError("Không tìm thấy dòng dữ liệu nào trong file lịch hỗ trợ.")
    return rows


# ---------------------------------------------------------------------------
# LỊCH PHÂN CA (nhiều ngày x 6 ca/ngày) + TỰ ĐỘNG PHÂN LINE
# ---------------------------------------------------------------------------

# Mã nhân viên -> tên ngắn dùng trong quy tắc phân line
MA_NV_TO_TEN_NGAN = {
    "112006": "Mi",
    "158278": "Quyên",
    "198434": "Sang",
    "214480": "Thi",
    "237593": "Ánh",
    "249215": "Linh",
    # 237175 = Son (bảo vệ), 227216 = Quí (quản lý) -> KHÔNG đưa vào phân line
}
MA_NV_LOAI_TRU = {"237175", "227216"}


def is_ca_schedule_file(input_path):
    """Kiểm tra nhanh: file có phải bảng lịch phân ca (nhiều ngày x Ca 1-6) không."""
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    row1 = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
    wb.close()
    return any("(" in v and ")" in v and "/" in v for v in row1)


def read_ca_schedule(input_path, nam_mac_dinh=None):
    """Đọc file lịch phân ca (giống bảng Quản Lý Phân Ca), trả về:
    {"2026-08-21": {"sang": ["Quyên","Sang","Thi","Linh"], "chieu": [...]}, ...}
    Chỉ lấy nhân viên có trong MA_NV_TO_TEN_NGAN (loại Quí, Son)."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    if nam_mac_dinh is None:
        nam_mac_dinh = datetime.now().year

    # Tìm các block ngày trên dòng 1: "T6 (21/08)" -> cột bắt đầu, ngày
    blocks = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and "(" in str(v) and ")" in str(v):
            inside = str(v).split("(")[1].split(")")[0].strip()
            parts = inside.replace("-", "/").split("/")
            if len(parts) == 2:
                day, month = int(parts[0]), int(parts[1])
                ngay_str = f"{nam_mac_dinh:04d}-{month:02d}-{day:02d}"
                blocks.append((ngay_str, c))

    # Tìm dòng nhân viên: cột 2 = mã nhân viên
    ma_nv_rows = []
    for r in range(1, ws.max_row + 1):
        ma = ws.cell(row=r, column=2).value
        if ma and str(ma).strip() in MA_NV_TO_TEN_NGAN:
            ma_nv_rows.append((r, str(ma).strip()))

    result = {}
    for ngay_str, start_col in blocks:
        sang, chieu = [], []
        for r, ma in ma_nv_rows:
            ten_ngan = MA_NV_TO_TEN_NGAN[ma]
            ca_vals = [ws.cell(row=r, column=start_col + i).value for i in range(6)]
            if any(v is not None for v in ca_vals[0:3]):
                sang.append(ten_ngan)
            if any(v is not None for v in ca_vals[3:6]):
                chieu.append(ten_ngan)
        result[ngay_str] = {"sang": sang, "chieu": chieu}

    return result


# ---- Quy tắc phân line tự động (chốt cùng anh Quí) ----
def phan_line_assign(names_in_ca, rotation_picker):
    """Áp quy tắc phân line cho 1 ca (list tên ngắn có mặt ca đó).
    rotation_picker(candidates_sorted_list) -> trả về 1 tên được chọn (để xoay vòng công bằng,
    do storage.py quản lý trạng thái xoay vòng).
    Trả về dict {"THU_NGAN": [...], "FRESH": [...], "FMCG": [...]}.
    """
    remaining = set(names_in_ca)
    result = {"THU_NGAN": [], "FRESH": [], "FMCG": []}

    # FMCG: Linh > Sang > Thi
    for uu_tien in ("Linh", "Sang", "Thi"):
        if uu_tien in remaining:
            result["FMCG"].append(uu_tien)
            remaining.discard(uu_tien)
            break

    # Mi luôn cố định THU NGÂN
    if "Mi" in remaining:
        result["THU_NGAN"].append("Mi")
        remaining.discard("Mi")

    # Sang chỉ đứng THU NGÂN hoặc FMCG -> nếu còn lại (chưa bị chọn FMCG) thì vào THU NGÂN
    if "Sang" in remaining:
        result["THU_NGAN"].append("Sang")
        remaining.discard("Sang")

    so_nguoi = len(names_in_ca)
    target_thu_ngan = 3 if so_nguoi >= 5 else 2

    # Quyên ưu tiên đứng THU NGÂN hơn FRESH
    if "Quyên" in remaining and len(result["THU_NGAN"]) < target_thu_ngan:
        result["THU_NGAN"].append("Quyên")
        remaining.discard("Quyên")

    # Các slot THU NGÂN còn thiếu -> xoay vòng trong số còn lại
    while len(result["THU_NGAN"]) < target_thu_ngan and remaining:
        candidates = sorted(remaining)
        pick = rotation_picker(candidates)
        result["THU_NGAN"].append(pick)
        remaining.discard(pick)

    # Còn lại -> FRESH
    for p in sorted(remaining):
        result["FRESH"].append(p)

    return result


TEN_NGAN_TO_MA_NV = {v: k for k, v in MA_NV_TO_TEN_NGAN.items()}

NOI_DUNG_CA_MAC_DINH = {
    "THU_NGAN": {
        "sang": ["NẤM 150k", "C2 8 lốc", "Bánh trung thu 6 cái"],
        "chieu": ["NẤM 300k", "C2 12 lốc", "Bánh trung thu 6 cái"],
    },
    "FRESH": {
        "sang": ["=> BC LỰA TRÁI CÂY TRƯỚC 8h"],
        "chieu": ["HẾT THỊT - CÁ GIẢM GIÁ", "HẾT RAU LÁ"],
    },
    "FMCG": {
        "sang": ["Xử lí hàng FMCG ( nếu có )", "Bắn kệ , châm hàng", "Quét lau FMCG + Kho"],
        "chieu": ["Xử lí hàng FMCG ( nếu có )", "Dọn dẹp kho bãi gọn ràng có lối đi", "BC FMCG"],
    },
}
