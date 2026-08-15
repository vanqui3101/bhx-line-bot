"""
excel_report.py - Tạo file Excel báo cáo chi tiết doanh thu từng siêu thị,
so sánh với cùng ngày tháng trước.
(Giữ nguyên như code cũ, không thay đổi)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

FONT_NAME = "Arial"
NAVY = "1F4E78"
GREEN = "1DB446"
RED = "E0483F"
GRAY = "999999"


def _total_dt(rec):
    return (rec.get("dt_offline") or 0) + (rec.get("dt_online") or 0)


def _total_bill(rec):
    return (rec.get("bill_offline") or 0) + (rec.get("bill_online") or 0)


def build_detail_excel(latest_date, latest_records, prev_date, prev_records, output_path):
    total_dt_now = sum(_total_dt(r) for r in latest_records)
    total_bill_now = sum(_total_bill(r) for r in latest_records)
    avg_bill_now = (total_dt_now / total_bill_now) if total_bill_now else 0

    prev_by_store = {r["ma_st"]: _total_dt(r) for r in prev_records} if prev_records else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiết doanh thu"

    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    title_font = Font(name=FONT_NAME, bold=True, size=14, color=NAVY)
    label_font = Font(name=FONT_NAME, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    date_display = datetime.strptime(latest_date, "%Y-%m-%d").strftime("%d/%m/%Y")
    prev_display = datetime.strptime(prev_date, "%Y-%m-%d").strftime("%d/%m/%Y") if prev_date else "Chưa có dữ liệu"

    ws.merge_cells('A1:G1')
    ws['A1'] = f"BÁO CÁO DOANH THU CHI TIẾT - {date_display}"
    ws['A1'].font = title_font

    ws.merge_cells('A2:G2')
    ws['A2'] = f"So với cùng ngày tháng trước: {prev_display}"
    ws['A2'].font = Font(name=FONT_NAME, size=10, color="666666")

    ws.merge_cells('A3:G3')
    avg_bill_str = f"{avg_bill_now:,.0f}".replace(",", ".")
    ws['A3'] = f"Giá trị bill trung bình toàn hệ thống (ngày {date_display}): {avg_bill_str} đ"
    ws['A3'].font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)

    r = 5
    headers = ["Siêu thị", "DT Offline", "DT Online", "Tổng DT (đến hiện tại)", "DT tháng trước", "Chênh lệch", "%"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1

    rows_sorted = sorted(latest_records, key=lambda x: -_total_dt(x))
    for rec in rows_sorted:
        dt_off = rec.get("dt_offline") or 0
        dt_on = rec.get("dt_online") or 0
        total_now = dt_off + dt_on
        prev_val = prev_by_store.get(rec["ma_st"])
        delta = (total_now - prev_val) if prev_val is not None else None
        pct = (delta / prev_val) if (prev_val) else None

        ws.cell(row=r, column=1, value=rec.get("ten_st") or rec.get("ma_st")).font = label_font

        c2 = ws.cell(row=r, column=2, value=dt_off)
        c2.font = label_font
        c2.number_format = '#,##0" đ"'

        c3 = ws.cell(row=r, column=3, value=dt_on)
        c3.font = label_font
        c3.number_format = '#,##0" đ"'

        c4 = ws.cell(row=r, column=4, value=total_now)
        c4.font = Font(name=FONT_NAME, bold=True, size=10)
        c4.number_format = '#,##0" đ"'

        c5 = ws.cell(row=r, column=5, value=prev_val if prev_val is not None else "—")
        c5.font = label_font
        if prev_val is not None:
            c5.number_format = '#,##0" đ"'

        c6 = ws.cell(row=r, column=6, value=delta if delta is not None else "—")
        c6.font = Font(name=FONT_NAME, size=10, color=(GREEN if (delta or 0) >= 0 else RED) if delta is not None else GRAY)
        if delta is not None:
            c6.number_format = '+#,##0" đ";-#,##0" đ"'

        c7 = ws.cell(row=r, column=7, value=pct if pct is not None else "—")
        c7.font = Font(name=FONT_NAME, size=10, bold=True, color=(GREEN if (pct or 0) >= 0 else RED) if pct is not None else GRAY)
        if pct is not None:
            c7.number_format = '+0.0%;-0.0%'

        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border
        r += 1

    widths = [26, 15, 15, 18, 16, 15, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
